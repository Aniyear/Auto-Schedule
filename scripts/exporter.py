import json
import pandas as pd
from pathlib import Path
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
import os

def is_physical_education(course_name):
    name = str(course_name).lower().strip()
    return "physical education" in name or name == "pe"

def export_schedule(chromosome, json_path, excel_path):
    # Assign Gym or Online for appropriate events
    for gene in chromosome.genes:
        if is_physical_education(gene.course):
            gene.room = "Gym"
        if getattr(gene, "delivery_mode", "offline") == "online" and gene.type.lower() == "lecture":
            gene.room = "Online"
    export_to_json(chromosome, json_path)
    export_to_excel(chromosome, excel_path)

def export_to_json(chromosome, path):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    data = chromosome.to_json()
    for group_name in data:
        for entry in data[group_name]:
            course_name = entry.get("Course", "")
            if is_physical_education(course_name):
                entry["Room"] = "Gym"
            if entry.get("delivery_mode", "offline") == "online" and entry.get("type", "").lower() == "lecture":
                entry["Room"] = "Online"
            if "Instructor" in entry:
                del entry["Instructor"]
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def export_to_excel(chromosome, path):
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")

    rows = []
    for gene in sorted(chromosome.genes, key=lambda g: (g.group, g.day, g.time)):
        if getattr(gene, "delivery_mode", "offline") == "online" and gene.type.lower() == "lecture":
            room = "Online"
        elif is_physical_education(gene.course):
            room = "Gym"
        else:
            room = gene.room
        rows.append({
            "Group": gene.group,
            "Day of the week": gene.day,
            "Time": gene.time,
            "Discipline": gene.course,
            "Classroom": room,
            "Type": gene.type,
            "Lecturer": getattr(gene, "instructor", "")
        })

    df = pd.DataFrame(rows)
    writer = pd.ExcelWriter(path, engine="openpyxl")
    Path(path).parent.mkdir(parents=True, exist_ok=True)

    DAY_ORDER = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat"]

    for group, group_df in df.groupby("Group"):
        group_df["Day of the week"] = pd.Categorical(group_df["Day of the week"], categories=DAY_ORDER, ordered=True)
        sheet_df = group_df.sort_values(by=["Day of the week", "Time"])
        sheet_df.drop(columns=["Group"], inplace=True)
        sheet_name = group[:31]
        sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)

    writer.close()

    # === APPLY FORMATTING ===
    wb = load_workbook(path)
    for sheet in wb.sheetnames:
        ws = wb[sheet]

        # Add group title
        ws.insert_rows(1)
        ws.insert_rows(2)
        ws["A1"] = f"Group {sheet}"
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
        ws["A1"].font = Font(name="Times New Roman", size=12, bold=True)
        ws["A1"].alignment = Alignment(horizontal="center")

        # Apply header style
        for cell in ws[3]:
            cell.font = Font(name="Times New Roman", bold=True, size=7)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = header_fill
            cell.border = thin_border

        # Group same Day values and format all cells
        current_day = None
        start_row = 4
        for row_idx in range(4, ws.max_row + 1):
            day_cell = ws.cell(row=row_idx, column=1)
            is_same_day = (day_cell.value == current_day)
            if not is_same_day:
                if current_day is not None:
                    ws.merge_cells(start_row=start_row, start_column=1, end_row=row_idx - 1, end_column=1)
                    ws.cell(row=start_row, column=1).alignment = Alignment(horizontal='center', vertical='center')
                current_day = day_cell.value
                start_row = row_idx

            for col in range(1, 7):
                cell = ws.cell(row=row_idx, column=col)
                if col == 3 or col == 4:  # Discipline or Classroom
                    if isinstance(cell.value, str):
                        cell.value = cell.value.replace('/', '\n').replace(',', '\n')
                font_size = 11 if col == 1 or row_idx == 3 else 7
                cell.font = Font(name="Times New Roman", size=font_size)
                cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="top")
                cell.border = thin_border

        # Merge final day block
        if current_day and start_row < ws.max_row:
            ws.merge_cells(start_row=start_row, start_column=1, end_row=ws.max_row, end_column=1)
            ws.cell(row=start_row, column=1).alignment = Alignment(horizontal='center', vertical='center')

        # Auto-width
        for col in ws.columns:
            max_length = max((len(str(cell.value)) for cell in col if cell.value), default=0)
            adjusted_length = min(max_length + 2, 30)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = adjusted_length

        # Add director signature line
        last_row = ws.max_row + 2
        ws.cell(row=last_row, column=6).value = "Director Of Academic Affairs Department __________"
        ws.cell(row=last_row, column=6).alignment = Alignment(horizontal="right")
        ws.cell(row=last_row, column=6).font = Font(name="Times New Roman", size=10)

    wb.save(path)
